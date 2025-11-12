import json
import os
import pdb
import time
import typing as t
import urllib
from io import BytesIO

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from tqdm import tqdm

from hrflow.hrflow import Hrflow

from ..storing import get_all_jobs, get_all_profiles
from .job import TEMPLATE_URL as JOB_TEMPLATE_URL
from .job import fill_work_sheet as fill_job_work_sheet
from .job import parsing_evaluator as job_parsing_evaluator
from .profile import TEMPLATE_URL as PROFILE_TEMPLATE_URL
from .profile import fill_work_sheet as fill_profile_work_sheet
from .profile import parsing_evaluator as profile_parsing_evaluator

STATISTICS_SHEET_NAME = "1. Statistics"


def load_workbook_from_url(url: str) -> Workbook:
    """
    Load an excel file from a url

    Args:
        url:                         <str>
                                     The url of the excel file

    Returns:
        <Workbook>
        The loaded workbook
    """
    file = urllib.request.urlopen(url).read()
    return load_workbook(filename=BytesIO(file))


def prepare_report_path(path: str) -> str:
    """
    Prepare the report path

    Args:
        path:                        <str>
                                     The path of the report
    """
    if os.path.isdir(path):
        return os.path.join(path, "parsing_evaluation.xlsx")
    if not path.endswith(".xlsx"):
        return f"{path}.xlsx"
    return path


def generate_parsing_evaluation_report(
    client: Hrflow,
    report_path: str,
    source_key: t.Optional[str] = None,
    board_key: t.Optional[str] = None,
    show_progress: bool = False,
):
    """
    Generate a parsing evaluation report

    If you want to generate a parsing evaluation report for jobs, you must
    provide the board_key.
    If you want to generate a parsing evaluation report for profiles, you must
    provide the source_key.

    board_key and source_key are optional string, you must provide only one of them.

    Args:
        client:                      <Hrflow>
                                     The client to use
        source_key:                  <Optional[str]>
                                     The source key where the profiles are
        board_key:                   <Optional[str]>
                                     The board key where the jobs are
        report_path:                 <str>
                                     The path of the report
                                     This can be a already existing directory where
                                     the report will be saved as parsing_evaluation.xlsx
                                     This can be directly the path of the report.
                                     If the path is not an excel file (xlsx),
                                     the report will be saved as {path}.xlsx
        show_progress:               <bool>
                                     Show the progress bar
    """

    if not source_key and not board_key:
        raise ValueError("You must provide either source_key or board_key")
    if source_key and board_key:
        raise ValueError("You must provide only one of source_key or board_key")

    if source_key:
        profile_list = get_all_profiles(client, source_key, show_progress)
        evaluation_list = profile_parsing_evaluator(profile_list, show_progress)

        work_book = load_workbook_from_url(PROFILE_TEMPLATE_URL)
        work_sheet = work_book[STATISTICS_SHEET_NAME]
        fill_profile_work_sheet(work_sheet, evaluation_list, show_progress)
    else:
        assert board_key is not None
        job_list = get_all_jobs(client, board_key, show_progress)
        evaluation_list = job_parsing_evaluator(job_list, show_progress)

        work_book = load_workbook_from_url(JOB_TEMPLATE_URL)
        work_sheet = work_book[STATISTICS_SHEET_NAME]
        fill_job_work_sheet(work_sheet, evaluation_list, show_progress)

    report_path = prepare_report_path(report_path)
    work_book.save(report_path)
    work_book.close()


def fetch_with_retries(url, headers, params, max_retries=3, delay=2):
    for attempt in range(max_retries):
        try:
            response = requests.get(url, headers=headers, params=params, timeout=30)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"Request error: {e}. Retrying in {delay} seconds...")
            time.sleep(delay)
            delay *= 2
    return None


def remove_illegal_chars(text):
    """Supprime les caractères illégaux pour éviter IllegalCharacterError."""
    if isinstance(text, str):
        ILLEGAL_CHARS = "".join(chr(i) for i in range(32) if i not in (9, 10, 13))
        return text.translate(str.maketrans("", "", ILLEGAL_CHARS))
    return text


def build_results_dataframe(profile_data, job_data, results):
    # 1. Convert results list to DataFrame
    df_results = pd.DataFrame(results)

    # 2. Convert job_data list to DataFrame
    df_jobs = pd.DataFrame(job_data)

    # 3. Convert profile_data dict to DataFrame
    df_profiles = pd.DataFrame(profile_data)

    # 4. Merge DataFrames
    merged_df = df_results.merge(df_jobs, on="Job Key", how="left").merge(
        df_profiles, on="Profile Key", how="left"
    )

    # 5. Setting Default value for Status and Rejected Reason
    merged_df["Status"] = "Neutral"
    merged_df["Rejected Reason"] = "Other"
    merged_df["Rank"] = ""
    merged_df["Versatility"] = "FALSE"

    # 6. Select and reorder columns to match your requirements
    df = merged_df[[
        "Profile Key",
        "Profile Reference",
        "Profile Link",
        "Profile Full Name",
        "Profile Summary",
        "Status",
        "Rejected Reason",
        "Rank",
        "Score",
        "Versatility",
        "Strengths",
        "Weaknesses",
        "Job Key",
        "Job Reference",
        "Job Link",
        "Job Title",
        "Job Summary",
    ]]

    # add change versatility to TRUE if strengths are present but weaknesses are empty
    df.loc[
        (df["Strengths"].notnull())
        & (df["Strengths"] != "")
        & ((df["Weaknesses"].isnull()) | (df["Weaknesses"] == "")),
        "Versatility",
    ] = "TRUE"

    # 7. Clean Unauthorized Chars
    df = df.map(remove_illegal_chars)

    return df


def clean_sheet_name(name):
    return name[:31].replace("/", "-").replace("\\", "-").replace(":", " ")


def save_to_excel(df, output_file):
    theme_structure = {
        "Profile": {
            "columns": ["Key", "Reference", "Link", "Full Name", "Summary"],
            "color": "FFCCCC",
        },
        "Recruiter Assessment": {
            "columns": ["Status", "Rejected Reason"],
            "color": "CCFFCC",
        },
        "AI Assessment": {"columns": ["Rank", "Score"], "color": "CCCCFF"},
        "AI Explainability": {
            "columns": ["Versatility", "Strengths", "Weaknesses"],
            "color": "FFF2CC",
        },
        "Job": {
            "columns": ["Key", "Reference", "Link", "Title", "Summary"],
            "color": "E0FFFF",
        },
    }

    if "Job Key" not in df.columns:
        raise KeyError(
            "'Job Key' column is missing. Do NOT rename columns before grouping."
        )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        processed_groups = []

        for job_key, job_group in df.groupby("Job Key"):
            job_group = job_group.copy()

            # Rename after grouping
            job_group.columns = [
                col.replace("Profile ", "").replace("Job ", "")
                for col in job_group.columns
            ]

            # Ranking
            job_group["Score"] = pd.to_numeric(job_group["Score"], errors="coerce")
            job_group["Rank"] = job_group["Score"].rank(ascending=False, method="min")
            job_group = job_group.sort_values(by="Score", ascending=False)

            job_title = (
                job_group["Title"].iloc[0] if "Title" in job_group.columns else "Job"
            )
            sheet_name = clean_sheet_name(str(job_title).strip())
            processed_groups.append((sheet_name, job_group))

        if not processed_groups:
            raise ValueError(
                "No job groups were processed. Check if 'Job Key' column is populated."
            )

        for sheet_name, job_group in processed_groups:
            job_group.to_excel(writer, sheet_name=sheet_name, index=False)

        # Formatting
        wb = writer.book
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Insert theme header
            sheet.insert_rows(1)
            col_index = 1
            for theme, props in theme_structure.items():
                span = len(props["columns"])
                cell = sheet.cell(row=1, column=col_index, value=theme)
                cell.fill = PatternFill(
                    start_color=props["color"],
                    end_color=props["color"],
                    fill_type="solid",
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)
                if span > 1:
                    sheet.merge_cells(
                        start_row=1,
                        start_column=col_index,
                        end_row=1,
                        end_column=col_index + span - 1,
                    )
                col_index += span

            # Column width & row height
            for col in range(1, sheet.max_column + 1):
                col_letter = get_column_letter(col)
                sheet.column_dimensions[col_letter].width = 14.2
            for row in range(1, sheet.max_row + 1):
                sheet.row_dimensions[row].height = 20

            # Format Score
            for col in sheet.iter_cols(min_row=2, max_row=2):
                if col[0].value == "Score":
                    score_col_index = col[0].column
                    for cell in sheet.iter_cols(
                        min_col=score_col_index,
                        max_col=score_col_index,
                        min_row=3,
                        max_row=sheet.max_row,
                    ):
                        for c in cell:
                            c.number_format = "0.00%"
                    break

            # Hyperlink for Link
            for col in sheet.iter_cols(min_row=2, max_row=2):
                if col[0].value == "Link":
                    col_index = col[0].column
                    for row in sheet.iter_rows(
                        min_row=3,
                        max_row=sheet.max_row,
                        min_col=col_index,
                        max_col=col_index,
                    ):
                        for cell in row:
                            if cell.value:
                                cell.hyperlink = cell.value
                                cell.style = "Hyperlink"
                                cell.font = Font(size=11)

    print(f"Excel file '{output_file}' created successfully!")


def generate_ranking_evaluation_report(
    client: Hrflow,
    source_key: str,
    board_key: str,
    scoring_algorithm: str,
    grading_algorithm: str,  # @param ["grader-hrflow-profiles", "grader-hrflow-profiles-recruiting", "grader-hrflow-jobs"]
    report_path: str,
    profile_tag_key: t.Optional[str] = "job_id",
    output_language: t.Optional[str] = "en",
    upskilling_compute_threshold: float = 0.7,
    show_progress: bool = False,
    max_retries: int = 3,
    delay: int = 2,
):
    """
    Applies the ranking on a set of profiles and jobs and generates an evaluation report.

    Prerequisites:
        - Grading, Scoring, Asking & Upskilling features must be enabled in the Hrflow team.
        - A set of profiles with known good matches to a set of jobs.
        - The profiles should be stored in a source with scoring enabled.
        - Each profile should have a tag, which the value of is a unique job identifier (the reference of the job in the board), indicating the job to which it should be graded.
        - A set of jobs stored in a board.

    Args:
        client:                      <Hrflow>
                                     The client to use
        source_key:                  <str>
                                     The source key where the profiles are stored
        board_key:                   <str>
                                     The board key where the jobs are stored
        scoring_algorithm:         <str>
                                     The scoring algorithm key to use
        grading_algorithm:          <str>
                                     The grading algorithm key to use
        profile_tag_key:             <str>
                                     The tag key in profiles that contains the job identifier
        output_language:            <Optional[str]>
                                     The language in which the report will be generated. default is 'en' (English).
        report_path:                 <str>
                                     The path of the report
        show_progress:               <bool>
                                     Show the progress bar
    """

    GRADING_URL = f"{client.api_url}/profile/grading"
    PROFILE_UPSKILLING_URL = f"{client.api_url}/profile/upskilling"

    profile_list = get_all_profiles(client, source_key, show_progress)
    job_list = get_all_jobs(client, board_key, show_progress)

    pdb.set_trace()
    # step 1: Generate summaries for each profile and job

    profile_data = []
    job_data = []

    if show_progress:
        profiles = tqdm(profile_list, desc="Enriching profiles")
    else:
        profiles = profile_list

    # Process profiles page by page
    for profile in profiles:

        first_name = profile.get("info", {}).get("first_name", "Unknown")
        last_name = profile.get("info", {}).get("last_name", "Unknown")
        profile_name = f"{first_name} {last_name}".strip()
        reference = profile["reference"]

        # Retrieve resume link
        resume_attachment = next(
            filter(lambda a: a["type"] == "resume", profile["attachments"]), {}
        )
        resume_link = resume_attachment.get("public_url", None)

        # Fetch profile summary
        lang_str = "FRENCH" if output_language == "fr" else "ENGLISH"

        summary_res = client.profile.asking.get(
            source_key=source_key,
            key=profile["key"],
            questions=[
                f"Generate a concise summary **IN {lang_str}** of the candidate's"
                " profile in the third person. Begin with their name, followed by an"
                " overview of their previous work experience and educational"
                " background. Do not mention the source name or reference any JSON"
                " objects. Ensure the summary is written in third person."
            ],
        )

        for attempt in range(max_retries):
            try:
                summary_res = client.profile.asking.get(
                    source_key=source_key,
                    key=profile["key"],
                    questions=[
                        f"Generate a concise summary **IN {lang_str}** of the"
                        " candidate's profile in the third person. Begin with their"
                        " name, followed by an overview of their previous work"
                        " experience and educational background. Do not mention the"
                        " source name or reference any JSON objects. Ensure the"
                        " summary is written in third person."
                    ],
                )
                break
            except Exception as e:
                print(f"Request error: {e}. Retrying in {delay} seconds...")
                time.sleep(delay)
                delay *= 2
        else:
            summary_res = None
        summary = (
            summary_res["data"][0] if summary_res and "data" in summary_res else ""
        )

        # Store the profile data
        profile_data.append({
            "Key": profile["key"],
            "Reference": reference,
            "Link": resume_link,
            "Full Name": profile_name,
            "Summary": summary,
        })

    if show_progress:
        jobs = tqdm(job_list, desc="Enriching jobs")
    else:
        jobs = job_list
    # Process jobs
    for job in jobs:

        lang_str = "FRENCH" if output_language == "fr" else "ENGLISH"

        # Second GET request to get the job summary and titles
        summary_res = client.job.asking.get(
            board_key=board_key,
            key=job["key"],
            questions=[
                (
                    f"Write a short summary in  **IN {lang_str}** for the job. Don't"
                    " mention board name nor JSON object. You'll start explaining work"
                    " environment, contract and duration, tasks and expected skills"
                    " for candidates."
                ),
            ],
        )

        for attempt in range(max_retries):
            try:
                summary_res = client.job.asking.get(
                    board_key=board_key,
                    key=job["key"],
                    questions=[
                        (
                            f"Write a short summary in  **IN {lang_str}** for the job."
                            " Don't mention board name nor JSON object. You'll start"
                            " explaining work environment, contract and duration,"
                            " tasks and expected skills for candidates."
                        ),
                    ],
                )
                break
            except Exception as e:
                print(f"Request error: {e}. Retrying in {delay} seconds...")
                time.sleep(delay)
                delay *= 2
        else:
            summary_res = None

        summary = (
            summary_res["data"][0] if summary_res and "data" in summary_res else ""
        )

        # Append the new job to the list
        job_data.append({
            "Key": job["key"],
            "Reference": job["reference"],
            "Link": job["url"],
            "Title": job["name"],
            "Summary": summary,
        })

    pdb.set_trace()
    # step 2: Grade profiles against jobs and generate report
    results = []

    if show_progress:
        job_data_iter = tqdm(job_data, desc="Processing Jobs")
    else:
        job_data_iter = job_data
    # Process jobs
    for job in job_data_iter:
        try:
            scoring_response = client.profile.scoring.list(
                source_keys=[source_key],
                board_key=board_key,
                job_key=job["Key"],
                algorithm_key=scoring_algorithm,
                use_algorithm=True,
                page=1,
                limit=1000,
                order_by="desc",
                sort_by="scoring",
                score_threshold=0,
                tags_included=[
                    [{"name": profile_tag_key, "value": job["Reference"]}]
                ],  # insert option to use profile_tag_key
            )

            profiles = scoring_response["data"].get("profiles", [])

            if show_progress:
                profiles = tqdm(
                    profiles,
                    desc=f"Grading profiles for job {job['reference']}",
                    leave=False,
                )
            # Keep track of already processed (job_key, profile_key) pairs
            existing_pairs = set(
                (result["Job Key"], result["Profile Key"]) for result in results
            )

            for i, profile in enumerate(profiles):
                # Skip grading if job_key & profile_key already exist
                if (job["Key"], profile["key"]) in existing_pairs:
                    continue

                # Fetch grading score
                grading_params = {
                    "source_key": source_key,
                    "profile_key": profile["key"],
                    "board_key": board_key,
                    "job_key": job["Key"],
                    "algorithm_key": grading_algorithm,
                }
                grading_res = fetch_with_retries(
                    url=GRADING_URL,
                    headers={
                        "accept": "application/json",
                        "X-API-KEY": client.auth_header["X-API-KEY"],
                        "X-USER-EMAIL": client.auth_header["X-USER-EMAIL"],
                    },
                    params=grading_params,
                    max_retries=max_retries,
                    delay=delay,
                )
                if (
                    not grading_res
                    or "data" not in grading_res
                    or "score" not in grading_res["data"]
                ):
                    print(
                        f'Skipping profile {profile["key"]} for job {job["key"]}: No'
                        " score found"
                    )
                    continue
                score = grading_res["data"]["score"]

                # Append new result
                results.append({
                    "Job Key": job["Key"],
                    "Profile Key": profile["key"],
                    "Score": score,
                })

                # Add to existing pairs to prevent re-processing
                existing_pairs.add((job["Key"], profile["key"]))

        except Exception as e:
            print(f'Unexpected error for job_key {job["Key"]}: {e}')

    print(" Processing completed successfully!")

    pdb.set_trace()
    # step 3: Compute upskilling

    if show_progress:
        results_iter = tqdm(
            results, desc="Computing upskilling for profile and job pairs"
        )
    else:
        results_iter = results
    for result in results_iter:
        if result["Score"] <= upskilling_compute_threshold:
            result["Strengths"] = ""
            result["Weaknesses"] = ""
            continue

        if all([
            "Strengths" in result,
            "Weaknesses" in result,
        ]):
            continue

        try:
            # Fetch Upskilling
            upskilling_params = {
                "board_key": board_key,
                "job_key": result["Job Key"],
                "source_key": source_key,
                "profile_key": result["Profile Key"],
                "score": result["Score"],
                "output_lang": output_language,
            }
            upskilling_res = fetch_with_retries(
                url=PROFILE_UPSKILLING_URL,
                headers={
                    "accept": "application/json",
                    "X-API-KEY": client.auth_header["X-API-KEY"],
                    "X-USER-EMAIL": client.auth_header["X-USER-EMAIL"],
                },
                params=upskilling_params,
                max_retries=max_retries,
                delay=delay,
            )
            if upskilling_res and "data" in upskilling_res:
                result["Strengths"] = "\n\n".join([
                    "\n".join([
                        strength["name"],
                        strength["description"],
                    ])
                    for strength in upskilling_res["data"]["strengths"]
                ])
                result["Weaknesses"] = "\n\n".join([
                    "\n".join([
                        strength["name"],
                        strength["description"],
                    ])
                    for strength in upskilling_res["data"]["weaknesses"]
                ])

        except Exception as e:
            print(f"Unexpected error for element {result}: {e}")
    print(" Processing completed successfully!")

    pdb.set_trace()
    # step 4: Build and save report
    profile_data = [
        {f"Profile {key}": value for key, value in data.items()}
        for data in profile_data
    ]
    job_data = [
        {f"Job {key}": value for key, value in data.items()} for data in job_data
    ]

    pdb.set_trace()
    df = build_results_dataframe(profile_data, job_data, results)

    save_to_excel(df, report_path)
