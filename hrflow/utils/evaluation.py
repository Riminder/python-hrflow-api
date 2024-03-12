import os
import typing as t
import urllib
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_interval
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel
from tqdm import tqdm

from .storing import get_all_profiles

TEMPLATE_URL = "https://riminder-documents-eu-2019-12-dev.s3.eu-west-1.amazonaws.com/evaluation/parsing-evaluation-template.xlsx"  # noqa: E501
STATISTICS_SHEET_NAME = "1. Statistics"
START_ROW_ID = 5

FILENAME_COLUMN_ID = "A"
RESUME_COLUMN_ID = "B"
PROFILE_COLUMN_ID = "C"

INFO_FIELD_LIST = (
    "score",
    "person",
    "first_name",
    "last_name",
    "phone",
    "email",
    "location",
    "summary",
    "driving_license",
)
INFO_START_COLUMN_ID, INFO_END_COLUMN_ID = ("D", "L")

EXPERIENCE_FIELD_LIST = (
    "score",
    "count",
    "title",
    "company",
    "start_date",
    "end_date",
    "location",
    "description",
    "skills",
    "tasks",
    "courses",
    "certifications",
)
EXPERIENCE_START_COLUMN_ID, EXPERIENCE_END_COLUMN_ID = ("M", "X")

EDUCATION_FIELD_LIST = (
    "score",
    "count",
    "title",
    "school",
    "start_date",
    "end_date",
    "location",
    "description",
    "skills",
    "tasks",
    "courses",
    "certifications",
)
EDUCATION_START_COLUMN_ID, EDUCATION_END_COLUMN_ID = ("Y", "AJ")

OTHER_FIELD_LIST = (
    "skills",
    "languages",
    "tasks",
    "courses",
    "certifications",
    "interests",
)
OTHER_START_COLUMN_ID, OTHER_END_COLUMN_ID = ("AK", "AP")


class InfoEvaluation(BaseModel):
    score: float
    person: float
    first_name: float
    last_name: float
    phone: float
    email: float
    location: float
    summary: float
    driving_license: float

    @staticmethod
    def from_profile(profile: t.Dict[str, t.Any]) -> "InfoEvaluation":
        info = profile["info"]
        first_name = 1 if info.get("first_name") else 0
        last_name = 1 if info.get("last_name") else 0
        phone = 1 if info.get("phone") else 0
        email = 1 if info.get("email") else 0
        location = 1 if info.get("location") else 0
        summary = 1 if info.get("summary") else 0
        driving_license = 1 if info.get("driving_license") else 0

        score = (
            first_name
            + last_name
            + phone
            + email
            + location
            + summary
            + driving_license
        )
        score /= 7
        return InfoEvaluation(
            score=score,
            person=1,
            first_name=first_name,
            last_name=last_name,
            phone=phone,
            email=email,
            location=location,
            summary=summary,
            driving_license=driving_license,
        )


class ExperienceEvaluation(BaseModel):
    score: float
    count: int
    title: float
    company: float
    start_date: float
    end_date: float
    location: float
    description: float
    skills: int
    tasks: int
    courses: int
    certifications: int

    @staticmethod
    def from_profile(profile: t.Dict[str, t.Any]) -> "ExperienceEvaluation":
        experiences = profile["experiences"]

        count = len(experiences)

        title = 0
        company = 0
        start_date = 0
        end_date = 0
        location = 0
        description = 0

        skills = 0
        tasks = 0
        courses = 0
        certifications = 0

        for experience in experiences:
            title += 1 if experience.get("title") else 0
            company += 1 if experience.get("company") else 0
            start_date += 1 if experience.get("date_start") else 0
            end_date += 1 if experience.get("date_end") else 0
            location += 1 if experience.get("location", {}).get("text") else 0
            description += 1 if experience.get("description") else 0

            skills += len(experience.get("skills", []))
            tasks += len(experience.get("tasks", []))
            courses += len(experience.get("courses", []))
            certifications += len(experience.get("certifications", []))

        if count > 0:
            title /= count
            company /= count
            start_date /= count
            end_date /= count
            location /= count
            description /= count

        score = title + company + start_date + end_date + location + description
        score /= 6

        return ExperienceEvaluation(
            score=score,
            count=count,
            title=title,
            company=company,
            start_date=start_date,
            end_date=end_date,
            location=location,
            description=description,
            skills=skills,
            tasks=tasks,
            courses=courses,
            certifications=certifications,
        )


class EducationEvaluation(BaseModel):
    score: float
    count: int
    title: float
    school: float
    start_date: float
    end_date: float
    location: float
    description: float
    skills: int
    tasks: int
    courses: int
    certifications: int

    @staticmethod
    def from_profile(profile: t.Dict[str, t.Any]) -> "EducationEvaluation":
        educations = profile["educations"]

        count = len(educations)

        title = 0
        school = 0
        start_date = 0
        end_date = 0
        location = 0
        description = 0

        skills = 0
        tasks = 0
        courses = 0
        certifications = 0

        for education in educations:
            title += 1 if education.get("title") else 0
            school += 1 if education.get("school") else 0
            start_date += 1 if education.get("date_start") else 0
            end_date += 1 if education.get("date_end") else 0
            location += 1 if education.get("location", {}).get("text") else 0
            description += 1 if education.get("description") else 0

            skills += len(education.get("skills", []))
            tasks += len(education.get("tasks", []))
            courses += len(education.get("courses", []))
            certifications += len(education.get("certifications", []))

        if count > 0:
            title /= count
            school /= count
            start_date /= count
            end_date /= count
            location /= count
            description /= count

        score = title + school + start_date + end_date + location + description
        score /= 6

        return EducationEvaluation(
            score=score,
            count=count,
            title=title,
            school=school,
            start_date=start_date,
            end_date=end_date,
            location=location,
            description=description,
            skills=skills,
            tasks=tasks,
            courses=courses,
            certifications=certifications,
        )


class OtherEvaluation(BaseModel):
    skills: int
    languages: int
    tasks: int
    courses: int
    certifications: int
    interests: int

    @staticmethod
    def from_profile(profile: t.Dict[str, t.Any]) -> "OtherEvaluation":
        skills = len(profile.get("skills", []))
        languages = len(profile.get("languages", []))
        tasks = len(profile.get("tasks", []))
        courses = len(profile.get("courses", []))
        certifications = len(profile.get("certifications", []))
        interests = len(profile.get("interests", []))

        return OtherEvaluation(
            skills=skills,
            languages=languages,
            tasks=tasks,
            courses=courses,
            certifications=certifications,
            interests=interests,
        )


class ProfileEvaluation(BaseModel):
    info: InfoEvaluation
    experience: ExperienceEvaluation
    education: EducationEvaluation
    other: OtherEvaluation

    filename: str
    resume_url: str
    profile_url: str

    @staticmethod
    def get_filename(profile: t.Dict[str, t.Any]) -> str:
        return profile["attachments"][0].get("original_file_name", "")

    @staticmethod
    def get_resume_url(profile: t.Dict[str, t.Any]) -> str:
        return profile["attachments"][0].get("public_url", "")

    @staticmethod
    def get_profile_url(profile: t.Dict[str, t.Any]) -> str:
        resume_url = ProfileEvaluation.get_resume_url(profile)
        base_url = resume_url.rsplit("/", 2)[0]
        return f"{base_url}/object.json"

    @staticmethod
    def from_profile(profile: t.Dict[str, t.Any]) -> "ProfileEvaluation":
        return ProfileEvaluation(
            info=InfoEvaluation.from_profile(profile),
            experience=ExperienceEvaluation.from_profile(profile),
            education=EducationEvaluation.from_profile(profile),
            other=OtherEvaluation.from_profile(profile),
            filename=ProfileEvaluation.get_filename(profile),
            resume_url=ProfileEvaluation.get_resume_url(profile),
            profile_url=ProfileEvaluation.get_profile_url(profile),
        )


def parsing_evaluator(
    profile_list: t.List[t.Dict[str, t.Any]], show_progress: bool = False
) -> t.List[ProfileEvaluation]:
    """
    Evaluate a list of profiles

    Args:
        profile_list:           <List[Dict[str, Any]>>
                                 List of profiles
        show_progress:          <bool>
                                 Show the progress bar

    Returns:
        <List[ProfileEvaluation]>:
        List of profile evaluations
    """
    if show_progress:
        profile_list = tqdm(profile_list, desc="Evaluating profiles")
    return [ProfileEvaluation.from_profile(profile) for profile in profile_list]


def load_workbook_from_url(url: str) -> Workbook:
    """
    Load an excel file from a url

    Args:
        url:                         <str>
                                     The url of the excel file

    Returns:
        <Workbook>
        The loaded workbook
    """
    file = urllib.request.urlopen(url).read()
    return load_workbook(filename=BytesIO(file))


def fill_metadata(
    work_sheet: Worksheet,
    profile_eval_list: t.List[ProfileEvaluation],
    show_progress: bool = False,
) -> None:
    """
    Fill the metadata of the profiles in the worksheet

    Args:
        work_sheet:                  <Worksheet>
                                     The worksheet to fill
        profile_eval_list:           <List[ProfileEvaluation]>
                                     The list of profile evaluations
        show_progress:               <bool>
                                     Show the progress bar
    """
    if show_progress:
        profile_eval_list = tqdm(profile_eval_list, desc="Filling meta-data")
    for row_id, profile_eval in enumerate(profile_eval_list, START_ROW_ID):
        work_sheet[f"{FILENAME_COLUMN_ID}{row_id}"].value = profile_eval.filename
        work_sheet[f"{RESUME_COLUMN_ID}{row_id}"].hyperlink = profile_eval.resume_url
        work_sheet[f"{PROFILE_COLUMN_ID}{row_id}"].hyperlink = profile_eval.profile_url


def fill_info(
    work_sheet: Worksheet,
    profile_eval_list: t.List[ProfileEvaluation],
    show_progress: bool = False,
) -> None:
    """
    Fill the info scores of the profiles in the worksheet

    Args:
        work_sheet:                  <Worksheet>
                                     The worksheet to fill
        profile_eval_list:           <List[ProfileEvaluation]>
                                     The list of profile evaluations
        show_progress:               <bool>
                                     Show the progress bar
    """
    colum_id_list = get_column_interval(INFO_START_COLUMN_ID, INFO_END_COLUMN_ID)
    if show_progress:
        profile_eval_list = tqdm(profile_eval_list, desc="Filling info scores")
    for row_id, profile_eval in enumerate(profile_eval_list, START_ROW_ID):
        for column_id, field in zip(colum_id_list, INFO_FIELD_LIST):
            work_sheet[f"{column_id}{row_id}"].value = getattr(profile_eval.info, field)


def fill_experience(
    work_sheet: Worksheet,
    profile_eval_list: t.List[ProfileEvaluation],
    show_progress: bool = False,
) -> None:
    """
    Fill the experience scores of the profiles in the worksheet

    Args:
        work_sheet:                  <Worksheet>
                                     The worksheet to fill
        profile_eval_list:           <List[ProfileEvaluation]>
                                     The list of profile evaluations
        show_progress:               <bool>
                                     Show the progress bar
    """
    colum_id_list = get_column_interval(
        EXPERIENCE_START_COLUMN_ID, EXPERIENCE_END_COLUMN_ID
    )
    if show_progress:
        profile_eval_list = tqdm(profile_eval_list, desc="Filling experience scores")
    for row_id, profile_eval in enumerate(profile_eval_list, START_ROW_ID):
        for column_id, field in zip(colum_id_list, EXPERIENCE_FIELD_LIST):
            work_sheet[f"{column_id}{row_id}"].value = getattr(
                profile_eval.experience, field
            )


def fill_education(
    work_sheet: Worksheet,
    profile_eval_list: t.List[ProfileEvaluation],
    show_progress: bool = False,
) -> None:
    """
    Fill the education scores of the profiles in the worksheet

    Args:
        work_sheet:                  <Worksheet>
                                     The worksheet to fill
        profile_eval_list:           <List[ProfileEvaluation]>
                                     The list of profile evaluations
        show_progress:               <bool>
                                     Show the progress bar
    """
    colum_id_list = get_column_interval(
        EDUCATION_START_COLUMN_ID, EDUCATION_END_COLUMN_ID
    )
    if show_progress:
        profile_eval_list = tqdm(profile_eval_list, desc="Filling education scores")
    for row_id, profile_eval in enumerate(profile_eval_list, START_ROW_ID):
        for column_id, field in zip(colum_id_list, EDUCATION_FIELD_LIST):
            work_sheet[f"{column_id}{row_id}"].value = getattr(
                profile_eval.education, field
            )


def fill_other(
    work_sheet: Worksheet,
    profile_eval_list: t.List[ProfileEvaluation],
    show_progress: bool = False,
) -> None:
    """
    Fill the other scores of the profiles in the worksheet

    Args:
        work_sheet:                  <Worksheet>
                                     The worksheet to fill
        profile_eval_list:           <List[ProfileEvaluation]>
                                     The list of profile evaluations
        show_progress:               <bool>
                                     Show the progress bar
    """
    colum_id_list = get_column_interval(OTHER_START_COLUMN_ID, OTHER_END_COLUMN_ID)
    if show_progress:
        profile_eval_list = tqdm(profile_eval_list, desc="Filling other scores")
    for row_id, profile_eval in enumerate(profile_eval_list, START_ROW_ID):
        for column_id, field in zip(colum_id_list, OTHER_FIELD_LIST):
            work_sheet[f"{column_id}{row_id}"].value = getattr(
                profile_eval.other, field
            )


def prepare_report_path(path: str) -> str:
    """
    Prepare the report path

    Args:
        path:                        <str>
                                     The path of the report
    """
    if os.path.isdir(path):
        return os.path.join(path, "parsing_evaluation.xlsx")
    if not path.endswith(".xlsx"):
        return f"{path}.xlsx"
    return path


def generate_parsing_evaluation_report(
    client: "Hrflow",  # noqa: F821
    source_key: str,
    report_path: str,
    show_progress: bool = False,
):
    """
    Generate a parsing evaluation report

    Args:
        client:                      <Hrflow>
                                     The client to use
        source_key:                  <str>
                                     The source key where the profiles are
        report_path:                 <str>
                                     The path of the report
                                     This can be a already existing directory where
                                     the report will be saved as parsing_evaluation.xlsx
                                     This can be directly the path of the report.
                                     If the path is not an excel file (xlsx),
                                     the report will be saved as {path}.xlsx
        show_progress:               <bool>
                                     Show the progress bar
    """
    profile_list = get_all_profiles(client, source_key, show_progress)
    evaluation_list = parsing_evaluator(profile_list, show_progress)

    work_book = load_workbook_from_url(TEMPLATE_URL)
    work_sheet = work_book[STATISTICS_SHEET_NAME]

    fill_metadata(work_sheet, evaluation_list, show_progress)
    fill_info(work_sheet, evaluation_list, show_progress)
    fill_experience(work_sheet, evaluation_list, show_progress)
    fill_education(work_sheet, evaluation_list, show_progress)
    fill_other(work_sheet, evaluation_list, show_progress)

    report_path = prepare_report_path(report_path)
    work_book.save(report_path)
    work_book.close()
